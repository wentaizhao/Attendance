import cv2
import numpy as np
import pandas as pd
import face_recognition
import keyboard
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date, datetime

print("----------Loading data----------")
try:
    attendance_list = pd.read_pickle("Attendance List.pkl")
except FileNotFoundError:
    print("\"Attendance List.pkl\" not found. Please run \"encode.exe.\"")
    input("Press Any Key to Close")
    raise SystemExit


print("----------Fetching Names----------")
people_names = attendance_list["Names"]
people_names = people_names.to_numpy().tolist()
print("----------Fetching Encodings----------")
encode_list_known = attendance_list["Encodings"]
encode_list_known = encode_list_known.to_numpy().tolist()


def is_integer(i):
    try:
        int(i)
        return True
    except ValueError:
        return False


def get_integer(i):
    while not is_integer(i):
        i = input('Invalid Number. Try Again: ')
    return int(i)


print("----------Creating Time Sheet----------")
sheet_name = date.today().strftime("%m-%d-%y")
try:
    wb = load_workbook("Attendance Times.xlsx")

    print()
    print("What sheet would you like to use?")
    print("\t0) New Sheet")
    for sheet_num in range(len(wb.sheetnames)):
        print(f"\t{sheet_num + 1}) {wb.sheetnames[sheet_num]}")
    choice = input("Enter number: ")
    choice = get_integer(choice)

    while choice not in range(len(wb.sheetnames) + 1):
        choice = input('Number out of Range. Try again: ')
        choice = get_integer(choice)

    if choice == 0:
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = sheet_name[0:8] + f" ({str(counter)})"
            counter += 1
        wb.create_sheet(sheet_name)
        print(f"Sheet Created: {sheet_name}")
    else:
        sheet_name = wb.sheetnames[int(choice) - 1]
        print(f"Sheet Loaded: {sheet_name}")
except FileNotFoundError:
    choice = 0
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name
    print(f"Sheet Created: {sheet_name}")
wb.save("Attendance Times.xlsx")

if choice == 0:
    print("----------Initializing Time Sheet----------")
    try:
        path1 = "SOURCE.xlsx"
        path2 = "Attendance Times.xlsx"

        wb1 = load_workbook(filename=path1)
        ws1 = wb1.worksheets[0]

        wb2 = load_workbook(filename=path2)
        ws2 = wb2[sheet_name]

        for w_row in ws1:
            for cell in w_row:
                ws2[cell.coordinate].value = cell.value
        ws2["B1"].value = "Time In"
        ws2["C1"].value = "Time Out"
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        wb2.save(path2)
    except FileNotFoundError:
        print("\"SOURCE.xlsx\" not found. Please run \"encode.exe.\"")
        input("Press Any Key to Close")
        raise SystemExit

    print("----------Initializing Log Sheet----------")
    try:
        wb = load_workbook("Attendance Log.xlsx")
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Log"
        for w_row in ws1:
            for cell in w_row:
                sheet[cell.coordinate].value = cell.value
        sheet.column_dimensions['A'].width = 20
        wb.save("Attendance Log.xlsx")
        wb = load_workbook("Attendance Log.xlsx")
    ws = wb.worksheets[0]
    ws.cell(row=1, column=ws.max_column+1).value = sheet_name
    ws.column_dimensions[get_column_letter(ws.max_column)].width = 12
    for col_cells in ws.iter_cols(min_col=ws.max_column, max_col=ws.max_column):
        for cell in col_cells:
            if cell.coordinate[1] != '1':
                cell.value = 0
    wb.save("Attendance Log.xlsx")
else:
    wb2 = load_workbook("Attendance Times.xlsx")
    ws2 = wb2[sheet_name]
    wb = load_workbook("Attendance Log.xlsx")
    ws = wb.worksheets[0]

try:
    with open("data.txt", 'r') as file:
        lines = file.readlines()
except FileNotFoundError:
    "\"data.txt\" not found."
    input("Press Any Key to Close")
    raise SystemExit

constants = ["Camera number: ", "Enable terminal (must be SINGLE key): ", "Exit: ", "Manual input: ",
             "Get history: ", "Get log: "]


def get_cmd(num):
    return (lines[num][len(constants[num]):]).strip()


CAMERA_NUMBER = int(get_cmd(0))
ENABLE_TERMINAL = get_cmd(1)
EXIT_KEY = get_cmd(2)
MANUAL_KEY = get_cmd(3)
HISTORY_KEY = get_cmd(4)
LOG_KEY = get_cmd(5)

WINDOW_SCALE_FACTOR = 2
TIME_SHEET = ws2
LOG_SHEET = ws
IN_COLUMN = 2
OUT_COLUMN = 3

history = []
try:
    with open(f"History\\{sheet_name}.txt", 'r') as history_file:
        temp_history = history_file.readlines()
    for line in temp_history:
        history.append(line.strip())
except FileNotFoundError:
    pass

log = []
target_col = 0
for row_cells in ws.iter_rows(min_row=1, max_row=1):
    for cell in row_cells:
        if cell.value == sheet_name:
            target_col = cell.coordinate[0]
target_col = ord(target_col) - 64
for col_cells in ws.iter_cols(min_col=target_col, max_col=target_col):
    for cell in col_cells:
        if cell.value == 1:
            name_cell = 'A' + cell.coordinate[1]
            log.append(LOG_SHEET[name_cell].value)

cap = cv2.VideoCapture(CAMERA_NUMBER, cv2.CAP_DSHOW)  # parameter sets correct camera


def time_add_name():
    if TIME_SHEET.cell(row=row, column=IN_COLUMN).value is None:
        time = datetime.now().strftime("%I:%M:%S %p")
        TIME_SHEET.cell(row=row, column=IN_COLUMN).value = time
        history.append(f"{time} {TIME_SHEET.cell(row=row, column=1).value} sign in")
        wb2.save("Attendance Times.xlsx")

    if TIME_SHEET.cell(row=row, column=OUT_COLUMN).value is None:
        sec_elapsed = datetime.now() - datetime.strptime(f"{sheet_name[0:8]} {TIME_SHEET.cell(row=row, column=IN_COLUMN).value}", "%m-%d-%y %I:%M:%S %p")
    else:
        sec_elapsed = datetime.now() - datetime.strptime(f"{sheet_name[0:8]} {TIME_SHEET.cell(row=row, column=OUT_COLUMN).value}", "%m-%d-%y %I:%M:%S %p")

    if sec_elapsed.total_seconds() > 300:
        time = datetime.now().strftime("%I:%M:%S %p")
        if TIME_SHEET.cell(row=row, column=OUT_COLUMN).value is None:
            history.append(f"{time} {TIME_SHEET.cell(row=row, column=1).value} sign out")
        else:
            history.append(f"{time} {TIME_SHEET.cell(row=row, column=1).value} update sign out")
        TIME_SHEET.cell(row=row, column=OUT_COLUMN).value = time
        wb2.save("Attendance Times.xlsx")
    # print(sec_elapsed.total_seconds())


def log_add_name():
    if LOG_SHEET.cell(row=row, column=LOG_SHEET.max_column).value == 0 or\
            LOG_SHEET.cell(row=row, column=LOG_SHEET.max_column).value is None:
        LOG_SHEET.cell(row=row, column=LOG_SHEET.max_column).value = 1
        log.append(TIME_SHEET.cell(row=row, column=1).value)
        log.sort()
        wb.save("Attendance Log.xlsx")


def print_list(my_list):
    for item in my_list:
        print(item)


if TIME_SHEET['F2'].value is None:
    print()
    leader = input("Enter who started this meeting: ")
    TIME_SHEET['F2'] = f"Meeting started by {leader}"
    wb2.save("Attendance Times.xlsx")
    print()

run_program = True
while run_program:
    success, img = cap.read()
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)  # small image to speed up
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)  # convert to rgb

    faceCurrFrame = face_recognition.face_locations(imgS)  # get face coordinates
    encodesCurrFrame = face_recognition.face_encodings(imgS, faceCurrFrame)  # encode faces

    for encodeFace, faceLoc in zip(encodesCurrFrame, faceCurrFrame):
        matches = face_recognition.compare_faces(encode_list_known, encodeFace, 0.4)  # compares face from cam with known faces
        faceDis = face_recognition.face_distance(encode_list_known, encodeFace)  # finds face distance between cam and known
        matchIndex = np.argmin(faceDis)  # get index of minimum in faceDis

        if matches[matchIndex]:  # if minimum is true
            name = people_names[matchIndex]
            row = matchIndex + 2
            time_add_name()
            log_add_name()

            name = name.upper()
            y1, x2, y2, x1 = faceLoc  # set coordinates of rectangle
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4  # scale rectangle back up
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)  # put rectangle on webcam image
            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)  # create solid area at bottom of rectangle
            cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)  # put name in area

    cv2.namedWindow("Webcam", cv2.WINDOW_NORMAL)
    width = img.shape[1] * WINDOW_SCALE_FACTOR
    height = img.shape[0] * WINDOW_SCALE_FACTOR
    cv2.resizeWindow("Webcam", width, height)
    cv2.imshow("Webcam", img)

    if keyboard.is_pressed(ENABLE_TERMINAL):
        cmd = input("Enter command or -1 to exit command terminal: ")
        print()
        while cmd != '-1':

            if cmd == EXIT_KEY:
                with open(f"History\\{sheet_name}.txt", 'w') as history_file:
                    for line in history:
                        history_file.write(line)
                        history_file.write('\n')
                # os.rename(f"{sheet_name}.txt", f"History\\{sheet_name}.txt")
                run_program = False
                break

            elif cmd == MANUAL_KEY:
                i_name = input("Enter Your Name: ")
                i_name = i_name.title()
                try:
                    row = people_names.index(i_name) + 2
                    time_add_name()
                    log_add_name()

                except ValueError:
                    i_name = '*' + i_name
                    row = TIME_SHEET.max_row + 1
                    TIME_SHEET.cell(row=row, column=1).value = i_name
                    row = LOG_SHEET.max_row + 1
                    LOG_SHEET.cell(row=row, column=1).value = i_name
                    time_add_name()
                    log_add_name()

                print("Name Added")

            elif cmd == HISTORY_KEY:
                print_list(history)

            elif cmd == LOG_KEY:
                print_list(log)

            else:
                print("Invalid Input. Try Again")

            print()
            cmd = input("Enter command or -1 to exit command terminal: ")
            print()

        print("Exiting command terminal")
        print()

    cv2.waitKey(1)

print("Exiting program")
cap.release()
cv2.destroyAllWindows()
input("Press Any Key to Close")



